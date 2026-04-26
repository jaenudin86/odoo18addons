import io
import json
from odoo import http
from odoo.http import request, content_disposition


class FinancialReportExportController(http.Controller):

    # ================================================================
    # HELPERS
    # ================================================================

    def _flatten_tree(self, nodes, rows=None, level=0):
        """Flatten hierarchy tree into list of rows with level info."""
        if rows is None:
            rows = []
        for node in nodes:
            row = dict(node)
            row['_level'] = level
            row.pop('children', None)
            rows.append(row)
            if node.get('children'):
                self._flatten_tree(node['children'], rows, level + 1)
        return rows

    def _get_xlsx_response(self, output, filename):
        response = request.make_response(
            output.getvalue(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition(filename)),
            ]
        )
        return response

    def _get_pdf_response(self, html_content, filename):
        """Generate PDF from HTML using wkhtmltopdf."""
        # Use Odoo's built-in IrActionsReport for wkhtmltopdf
        pdf_content = request.env['ir.actions.report']._run_wkhtmltopdf(
            [html_content],
            landscape=True,
        )
        response = request.make_response(
            pdf_content,
            headers=[
                ('Content-Type', 'application/pdf'),
                ('Content-Disposition', content_disposition(filename)),
            ]
        )
        return response

    def _build_pdf_html(self, title, subtitle, headers, rows, totals_row=None):
        """Build HTML table for PDF export."""
        css = """
        <style>
            body { font-family: Arial, sans-serif; font-size: 10px; margin: 20px; }
            h1 { font-size: 16px; margin-bottom: 2px; }
            h3 { font-size: 11px; color: #666; margin-top: 2px; }
            table { width: 100%; border-collapse: collapse; margin-top: 10px; }
            th { background: #2c3e50; color: #fff; padding: 6px 8px; text-align: left; font-size: 9px; text-transform: uppercase; }
            td { padding: 5px 8px; border-bottom: 1px solid #eee; font-size: 10px; }
            .text-right { text-align: right; }
            .level-0 { font-weight: bold; background: #f5f5f5; }
            .level-1 { padding-left: 20px; }
            .level-2 { padding-left: 40px; }
            .level-3 { padding-left: 60px; }
            .totals { font-weight: bold; background: #e9ecef; border-top: 2px solid #2c3e50; }
            .section-header { font-weight: bold; background: #2c3e50; color: #fff; }
            .section-header td { color: #fff; padding: 8px; }
            .net-row { font-weight: bold; background: #1a252f; color: #fff; font-size: 12px; }
            .net-row td { color: #fff; padding: 8px; }
        </style>
        """
        html = f"<html><head>{css}</head><body>"
        html += f"<h1>{title}</h1><h3>{subtitle}</h3>"
        html += "<table><thead><tr>"
        for h in headers:
            align = ' class="text-right"' if h.get('right') else ''
            html += f"<th{align}>{h['label']}</th>"
        html += "</tr></thead><tbody>"
        for row in rows:
            css_class = row.get('_css', '')
            html += f'<tr class="{css_class}">'
            for h in headers:
                val = row.get(h['key'], '')
                align = ' class="text-right"' if h.get('right') else ''
                level = row.get('_level', 0)
                indent = ''
                if h['key'] == 'name' and level > 0:
                    indent = f' style="padding-left:{level * 20 + 8}px;"'
                    align = ''
                if isinstance(val, (int, float)):
                    val = f"{val:,.2f}"
                html += f"<td{align}{indent}>{val}</td>"
            html += "</tr>"
        if totals_row:
            html += '<tr class="totals">'
            for h in headers:
                val = totals_row.get(h['key'], '')
                align = ' class="text-right"' if h.get('right') else ''
                if isinstance(val, (int, float)):
                    val = f"{val:,.2f}"
                html += f"<td{align}>{val}</td>"
            html += "</tr>"
        html += "</tbody></table></body></html>"
        return html

    # ================================================================
    # TRIAL BALANCE EXPORT
    # ================================================================

    @http.route('/ac_financial_reports/export/trial_balance/xlsx', type='http', auth='user')
    def export_trial_balance_xlsx(self, date_from=None, date_to=None, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        data = request.env['account.account'].get_trial_balance_data(date_from=date_from, date_to=date_to)
        rows = self._flatten_tree(data['accounts'])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trial Balance"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        bold_font = Font(bold=True, size=10)
        num_fmt = '#,##0.00'
        thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = 'Trial Balance'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Period: {date_from or ''} to {date_to or ''}"
        ws['A2'].font = Font(size=10, color="666666")

        # Headers
        headers = ['Code', 'Account Name', 'Opening', 'Debit', 'Credit', 'Ending']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='right' if col > 2 else 'left')

        # Data
        for i, row in enumerate(rows):
            r = i + 5
            indent = '  ' * row['_level']
            ws.cell(row=r, column=1, value=row['code'])
            ws.cell(row=r, column=2, value=indent + row['name'])
            ws.cell(row=r, column=3, value=row['opening_balance']).number_format = num_fmt
            ws.cell(row=r, column=4, value=row['debit']).number_format = num_fmt
            ws.cell(row=r, column=5, value=row['credit']).number_format = num_fmt
            ws.cell(row=r, column=6, value=row['ending_balance']).number_format = num_fmt
            if row.get('is_parent'):
                for c in range(1, 7):
                    ws.cell(row=r, column=c).font = bold_font
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = thin_border

        # Totals
        r = len(rows) + 5
        totals_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        ws.cell(row=r, column=2, value="TOTAL").font = bold_font
        ws.cell(row=r, column=3, value=data['totals']['opening_balance']).number_format = num_fmt
        ws.cell(row=r, column=4, value=data['totals']['debit']).number_format = num_fmt
        ws.cell(row=r, column=5, value=data['totals']['credit']).number_format = num_fmt
        ws.cell(row=r, column=6, value=data['totals']['ending_balance']).number_format = num_fmt
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = totals_fill
            ws.cell(row=r, column=c).font = bold_font

        # Column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 40
        for col in ['C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 18

        output = io.BytesIO()
        wb.save(output)
        return self._get_xlsx_response(output, f"trial_balance_{date_from}_{date_to}.xlsx")

    @http.route('/ac_financial_reports/export/trial_balance/pdf', type='http', auth='user')
    def export_trial_balance_pdf(self, date_from=None, date_to=None, **kwargs):
        data = request.env['account.account'].get_trial_balance_data(date_from=date_from, date_to=date_to)
        rows = self._flatten_tree(data['accounts'])
        headers = [
            {'key': 'code', 'label': 'Code'},
            {'key': 'name', 'label': 'Account'},
            {'key': 'opening_balance', 'label': 'Opening', 'right': True},
            {'key': 'debit', 'label': 'Debit', 'right': True},
            {'key': 'credit', 'label': 'Credit', 'right': True},
            {'key': 'ending_balance', 'label': 'Ending', 'right': True},
        ]
        totals = {'name': 'TOTAL', **data['totals']}
        html = self._build_pdf_html("Trial Balance", f"Period: {date_from} to {date_to}", headers, rows, totals)
        return self._get_pdf_response(html, f"trial_balance_{date_from}_{date_to}.pdf")

    # ================================================================
    # PROFIT & LOSS EXPORT
    # ================================================================

    @http.route('/ac_financial_reports/export/profit_loss/xlsx', type='http', auth='user')
    def export_profit_loss_xlsx(self, date_from=None, date_to=None, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

        data = request.env['account.account'].get_profit_loss_data(date_from=date_from, date_to=date_to)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Profit & Loss"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        section_fill_income = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        section_fill_expense = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        net_fill = PatternFill(start_color="1A252F", end_color="1A252F", fill_type="solid")
        bold_font = Font(bold=True, size=10)
        num_fmt = '#,##0.00'

        ws.merge_cells('A1:C1')
        ws['A1'] = 'Profit & Loss'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Period: {date_from or ''} to {date_to or ''}"

        headers = ['Code', 'Account Name', 'Balance']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        r = 5
        # Income section
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = section_fill_income
        ws.cell(row=r, column=2, value="INCOME").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=r, column=3, value=data['income']['total']).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=r, column=3).number_format = num_fmt
        r += 1

        for row in self._flatten_tree(data['income']['items']):
            indent = '  ' * row['_level']
            ws.cell(row=r, column=1, value=row['code'])
            ws.cell(row=r, column=2, value=indent + row['name'])
            ws.cell(row=r, column=3, value=row['balance']).number_format = num_fmt
            if row.get('is_parent'):
                for c in range(1, 4): ws.cell(row=r, column=c).font = bold_font
            r += 1

        # Expense section
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = section_fill_expense
        ws.cell(row=r, column=2, value="EXPENSES").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=r, column=3, value=data['expense']['total']).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=r, column=3).number_format = num_fmt
        r += 1

        for row in self._flatten_tree(data['expense']['items']):
            indent = '  ' * row['_level']
            ws.cell(row=r, column=1, value=row['code'])
            ws.cell(row=r, column=2, value=indent + row['name'])
            ws.cell(row=r, column=3, value=row['balance']).number_format = num_fmt
            if row.get('is_parent'):
                for c in range(1, 4): ws.cell(row=r, column=c).font = bold_font
            r += 1

        # Net profit
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = net_fill
        ws.cell(row=r, column=2, value="NET PROFIT / (LOSS)").font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(row=r, column=3, value=data['net_profit']).font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(row=r, column=3).number_format = num_fmt

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 20

        output = io.BytesIO()
        wb.save(output)
        return self._get_xlsx_response(output, f"profit_loss_{date_from}_{date_to}.xlsx")

    @http.route('/ac_financial_reports/export/profit_loss/pdf', type='http', auth='user')
    def export_profit_loss_pdf(self, date_from=None, date_to=None, **kwargs):
        data = request.env['account.account'].get_profit_loss_data(date_from=date_from, date_to=date_to)
        headers = [{'key': 'code', 'label': 'Code'}, {'key': 'name', 'label': 'Account'}, {'key': 'balance', 'label': 'Balance', 'right': True}]
        rows = []
        rows.append({'code': '', 'name': 'INCOME', 'balance': data['income']['total'], '_level': 0, '_css': 'section-header'})
        rows.extend(self._flatten_tree(data['income']['items']))
        rows.append({'code': '', 'name': 'EXPENSES', 'balance': data['expense']['total'], '_level': 0, '_css': 'section-header'})
        rows.extend(self._flatten_tree(data['expense']['items']))
        rows.append({'code': '', 'name': 'NET PROFIT / (LOSS)', 'balance': data['net_profit'], '_level': 0, '_css': 'net-row'})
        html = self._build_pdf_html("Profit & Loss", f"Period: {date_from} to {date_to}", headers, rows)
        return self._get_pdf_response(html, f"profit_loss_{date_from}_{date_to}.pdf")

    # ================================================================
    # BALANCE SHEET EXPORT
    # ================================================================

    @http.route('/ac_financial_reports/export/balance_sheet/xlsx', type='http', auth='user')
    def export_balance_sheet_xlsx(self, date_to=None, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        data = request.env['account.account'].get_balance_sheet_data(date_to=date_to)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        bold_font = Font(bold=True, size=10)
        num_fmt = '#,##0.00'

        ws.merge_cells('A1:C1')
        ws['A1'] = 'Balance Sheet'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"As of: {date_to or ''}"

        for col, h in enumerate(['Code', 'Account Name', 'Balance'], 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        r = 5
        sections = [
            ('ASSETS', data['assets'], "2563EB"),
            ('LIABILITIES', data['liabilities'], "DC2626"),
            ('EQUITY', data['equity'], "059669"),
        ]
        for section_name, section_data, color in sections:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for c in range(1, 4): ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=2, value=section_name).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=r, column=3, value=section_data['total']).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=r, column=3).number_format = num_fmt
            r += 1
            for row in self._flatten_tree(section_data['items']):
                indent = '  ' * row['_level']
                ws.cell(row=r, column=1, value=row['code'])
                ws.cell(row=r, column=2, value=indent + row['name'])
                ws.cell(row=r, column=3, value=row['balance']).number_format = num_fmt
                if row.get('is_parent'):
                    for c in range(1, 4): ws.cell(row=r, column=c).font = bold_font
                r += 1

        # Unallocated earnings
        ws.cell(row=r, column=2, value="Unallocated Earnings")
        ws.cell(row=r, column=3, value=data['unallocated_earnings']).number_format = num_fmt
        r += 1

        # Totals
        net_fill = PatternFill(start_color="1A252F", end_color="1A252F", fill_type="solid")
        for c in range(1, 4): ws.cell(row=r, column=c).fill = net_fill
        ws.cell(row=r, column=2, value="TOTAL LIABILITIES + EQUITY").font = Font(bold=True, color="FFFFFF", size=11)
        ws.cell(row=r, column=3, value=data['total_liabilities_equity']).font = Font(bold=True, color="FFFFFF", size=11)
        ws.cell(row=r, column=3).number_format = num_fmt

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 20

        output = io.BytesIO()
        wb.save(output)
        return self._get_xlsx_response(output, f"balance_sheet_{date_to}.xlsx")

    @http.route('/ac_financial_reports/export/balance_sheet/pdf', type='http', auth='user')
    def export_balance_sheet_pdf(self, date_to=None, **kwargs):
        data = request.env['account.account'].get_balance_sheet_data(date_to=date_to)
        headers = [{'key': 'code', 'label': 'Code'}, {'key': 'name', 'label': 'Account'}, {'key': 'balance', 'label': 'Balance', 'right': True}]
        rows = []
        for section_name, section_data in [('ASSETS', data['assets']), ('LIABILITIES', data['liabilities']), ('EQUITY', data['equity'])]:
            rows.append({'code': '', 'name': section_name, 'balance': section_data['total'], '_level': 0, '_css': 'section-header'})
            rows.extend(self._flatten_tree(section_data['items']))
        rows.append({'code': '', 'name': 'Unallocated Earnings', 'balance': data['unallocated_earnings'], '_level': 0, '_css': ''})
        rows.append({'code': '', 'name': 'TOTAL LIABILITIES + EQUITY', 'balance': data['total_liabilities_equity'], '_level': 0, '_css': 'net-row'})
        html = self._build_pdf_html("Balance Sheet", f"As of: {date_to}", headers, rows)
        return self._get_pdf_response(html, f"balance_sheet_{date_to}.pdf")

    # ================================================================
    # GENERAL LEDGER EXPORT
    # ================================================================

    @http.route('/ac_financial_reports/export/general_ledger/xlsx', type='http', auth='user')
    def export_general_ledger_xlsx(self, date_from=None, date_to=None, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        data = request.env['account.account'].get_general_ledger_data(date_from=date_from, date_to=date_to)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "General Ledger"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        acc_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        bold_font = Font(bold=True, size=10)
        num_fmt = '#,##0.00'

        ws.merge_cells('A1:G1')
        ws['A1'] = 'General Ledger'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Period: {date_from or ''} to {date_to or ''}"

        r = 4
        for acc in data['accounts']:
            # Account header
            for c in range(1, 8): ws.cell(row=r, column=c).fill = acc_fill
            ws.cell(row=r, column=1, value=acc['code']).font = bold_font
            ws.cell(row=r, column=2, value=acc['name']).font = bold_font
            ws.cell(row=r, column=7, value=acc['ending_balance']).font = bold_font
            ws.cell(row=r, column=7).number_format = num_fmt
            r += 1

            # Column headers
            for col, h in enumerate(['Date', 'Reference', 'Partner', 'Label', 'Debit', 'Credit', 'Balance'], 1):
                cell = ws.cell(row=r, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
            r += 1

            # Opening
            ws.cell(row=r, column=4, value="Opening Balance").font = Font(italic=True)
            ws.cell(row=r, column=7, value=acc['opening_balance']).number_format = num_fmt
            r += 1

            # Entries
            for entry in acc['entries']:
                ws.cell(row=r, column=1, value=entry['date'])
                ws.cell(row=r, column=2, value=entry['move_name'])
                ws.cell(row=r, column=3, value=entry['partner'])
                ws.cell(row=r, column=4, value=entry['label'])
                ws.cell(row=r, column=5, value=entry['debit']).number_format = num_fmt
                ws.cell(row=r, column=6, value=entry['credit']).number_format = num_fmt
                ws.cell(row=r, column=7, value=entry['balance']).number_format = num_fmt
                r += 1

            # Subtotal
            ws.cell(row=r, column=4, value="Ending Balance").font = bold_font
            ws.cell(row=r, column=5, value=acc['total_debit']).font = bold_font
            ws.cell(row=r, column=5).number_format = num_fmt
            ws.cell(row=r, column=6, value=acc['total_credit']).font = bold_font
            ws.cell(row=r, column=6).number_format = num_fmt
            ws.cell(row=r, column=7, value=acc['ending_balance']).font = bold_font
            ws.cell(row=r, column=7).number_format = num_fmt
            r += 2  # blank row between accounts

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        for col in ['E', 'F', 'G']:
            ws.column_dimensions[col].width = 16

        output = io.BytesIO()
        wb.save(output)
        return self._get_xlsx_response(output, f"general_ledger_{date_from}_{date_to}.xlsx")

    @http.route('/ac_financial_reports/export/general_ledger/pdf', type='http', auth='user')
    def export_general_ledger_pdf(self, date_from=None, date_to=None, **kwargs):
        data = request.env['account.account'].get_general_ledger_data(date_from=date_from, date_to=date_to)
        headers = [
            {'key': 'date', 'label': 'Date'}, {'key': 'move_name', 'label': 'Reference'},
            {'key': 'partner', 'label': 'Partner'}, {'key': 'label', 'label': 'Label'},
            {'key': 'debit', 'label': 'Debit', 'right': True}, {'key': 'credit', 'label': 'Credit', 'right': True},
            {'key': 'balance', 'label': 'Balance', 'right': True},
        ]
        rows = []
        for acc in data['accounts']:
            rows.append({'date': acc['code'], 'move_name': '', 'partner': '', 'label': acc['name'], 'debit': '', 'credit': '', 'balance': acc['ending_balance'], '_level': 0, '_css': 'level-0'})
            rows.append({'date': '', 'move_name': '', 'partner': '', 'label': 'Opening Balance', 'debit': '', 'credit': '', 'balance': acc['opening_balance'], '_level': 0, '_css': ''})
            for entry in acc['entries']:
                rows.append({**entry, '_level': 0, '_css': ''})
            rows.append({'date': '', 'move_name': '', 'partner': '', 'label': 'Ending Balance', 'debit': acc['total_debit'], 'credit': acc['total_credit'], 'balance': acc['ending_balance'], '_level': 0, '_css': 'totals'})
        html = self._build_pdf_html("General Ledger", f"Period: {date_from} to {date_to}", headers, rows)
        return self._get_pdf_response(html, f"general_ledger_{date_from}_{date_to}.pdf")

    # ================================================================
    # CASH FLOW EXPORT
    # ================================================================

    @http.route('/ac_financial_reports/export/cash_flow/xlsx', type='http', auth='user')
    def export_cash_flow_xlsx(self, date_from=None, date_to=None, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        data = request.env['account.account'].get_cash_flow_data(date_from=date_from, date_to=date_to)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cash Flow"

        bold_font = Font(bold=True, size=10)
        section_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        total_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        num_fmt = '#,##0.00'

        ws.merge_cells('A1:B1')
        ws['A1'] = 'Cash Flow Statement (Indirect Method)'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Period: {date_from or ''} to {date_to or ''}"

        r = 4
        def write_section(title, items, total, total_label):
            nonlocal r
            for c in [1, 2]:
                ws.cell(row=r, column=c).fill = section_fill
            ws.cell(row=r, column=1, value=title).font = Font(bold=True, color="FFFFFF")
            r += 1
            for item in items:
                ws.cell(row=r, column=1, value='  ' + item['name'])
                ws.cell(row=r, column=2, value=item['amount']).number_format = num_fmt
                r += 1
            for c in [1, 2]:
                ws.cell(row=r, column=c).fill = total_fill
            ws.cell(row=r, column=1, value=total_label).font = bold_font
            ws.cell(row=r, column=2, value=total).font = bold_font
            ws.cell(row=r, column=2).number_format = num_fmt
            r += 2

        # Operating
        ws.cell(row=r, column=1, value='Net Income').font = bold_font
        ws.cell(row=r, column=2, value=data['operating']['net_income']).number_format = num_fmt
        r += 1
        all_operating = data['operating']['adjustments'] + data['operating']['working_capital']
        write_section("OPERATING ACTIVITIES", all_operating, data['operating']['total'], "Net Cash from Operating")

        # Investing
        write_section("INVESTING ACTIVITIES", data['investing']['items'], data['investing']['total'], "Net Cash from Investing")

        # Financing
        write_section("FINANCING ACTIVITIES", data['financing']['items'], data['financing']['total'], "Net Cash from Financing")

        # Summary
        ws.cell(row=r, column=1, value='Opening Cash Balance')
        ws.cell(row=r, column=2, value=data['opening_cash']).number_format = num_fmt
        r += 1
        ws.cell(row=r, column=1, value='Net Change in Cash').font = bold_font
        ws.cell(row=r, column=2, value=data['net_cash_change']).number_format = num_fmt
        ws.cell(row=r, column=2).font = bold_font
        r += 1
        net_fill = PatternFill(start_color="1A252F", end_color="1A252F", fill_type="solid")
        for c in [1, 2]: ws.cell(row=r, column=c).fill = net_fill
        ws.cell(row=r, column=1, value='CLOSING CASH BALANCE').font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(row=r, column=2, value=data['closing_cash']).font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(row=r, column=2).number_format = num_fmt

        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 20

        output = io.BytesIO()
        wb.save(output)
        return self._get_xlsx_response(output, f"cash_flow_{date_from}_{date_to}.xlsx")

    @http.route('/ac_financial_reports/export/cash_flow/pdf', type='http', auth='user')
    def export_cash_flow_pdf(self, date_from=None, date_to=None, **kwargs):
        data = request.env['account.account'].get_cash_flow_data(date_from=date_from, date_to=date_to)
        headers = [{'key': 'name', 'label': 'Item'}, {'key': 'amount', 'label': 'Amount', 'right': True}]
        rows = []
        rows.append({'name': 'Net Income', 'amount': data['operating']['net_income'], '_level': 0, '_css': 'level-0'})
        rows.append({'name': 'OPERATING ACTIVITIES', 'amount': '', '_level': 0, '_css': 'section-header'})
        for item in data['operating']['adjustments'] + data['operating']['working_capital']:
            rows.append({'name': item['name'], 'amount': item['amount'], '_level': 1, '_css': ''})
        rows.append({'name': 'Net Cash from Operating', 'amount': data['operating']['total'], '_level': 0, '_css': 'totals'})
        rows.append({'name': 'INVESTING ACTIVITIES', 'amount': '', '_level': 0, '_css': 'section-header'})
        for item in data['investing']['items']:
            rows.append({'name': item['name'], 'amount': item['amount'], '_level': 1, '_css': ''})
        rows.append({'name': 'Net Cash from Investing', 'amount': data['investing']['total'], '_level': 0, '_css': 'totals'})
        rows.append({'name': 'FINANCING ACTIVITIES', 'amount': '', '_level': 0, '_css': 'section-header'})
        for item in data['financing']['items']:
            rows.append({'name': item['name'], 'amount': item['amount'], '_level': 1, '_css': ''})
        rows.append({'name': 'Net Cash from Financing', 'amount': data['financing']['total'], '_level': 0, '_css': 'totals'})
        rows.append({'name': 'Opening Cash', 'amount': data['opening_cash'], '_level': 0, '_css': ''})
        rows.append({'name': 'Net Change', 'amount': data['net_cash_change'], '_level': 0, '_css': 'level-0'})
        rows.append({'name': 'CLOSING CASH BALANCE', 'amount': data['closing_cash'], '_level': 0, '_css': 'net-row'})
        html = self._build_pdf_html("Cash Flow Statement (Indirect)", f"Period: {date_from} to {date_to}", headers, rows)
        return self._get_pdf_response(html, f"cash_flow_{date_from}_{date_to}.pdf")

    # ================================================================
    # AGING EXPORT
    # ================================================================

    @http.route('/ac_financial_reports/export/aging/xlsx', type='http', auth='user')
    def export_aging_xlsx(self, date_to=None, report_type='receivable', **kwargs):
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        data = request.env['account.account'].get_aging_data(date_to=date_to, report_type=report_type)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aging Report"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        bold_font = Font(bold=True, size=10)
        total_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        num_fmt = '#,##0.00'

        title = 'Aging Receivable' if report_type == 'receivable' else 'Aging Payable'
        ws.merge_cells('A1:G1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"As of: {date_to or ''}"

        headers = ['Partner', 'Current', '1-30', '31-60', '61-90', '>90', 'Total']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, p in enumerate(data['partners']):
            r = i + 5
            ws.cell(row=r, column=1, value=p['partner_name'])
            ws.cell(row=r, column=2, value=p['current']).number_format = num_fmt
            ws.cell(row=r, column=3, value=p['b1_30']).number_format = num_fmt
            ws.cell(row=r, column=4, value=p['b31_60']).number_format = num_fmt
            ws.cell(row=r, column=5, value=p['b61_90']).number_format = num_fmt
            ws.cell(row=r, column=6, value=p['b90_plus']).number_format = num_fmt
            ws.cell(row=r, column=7, value=p['total']).number_format = num_fmt
            ws.cell(row=r, column=7).font = bold_font

        # Totals
        r = len(data['partners']) + 5
        for c in range(1, 8): ws.cell(row=r, column=c).fill = total_fill
        ws.cell(row=r, column=1, value="TOTAL").font = bold_font
        for col, key in [(2, 'current'), (3, 'b1_30'), (4, 'b31_60'), (5, 'b61_90'), (6, 'b90_plus'), (7, 'total')]:
            ws.cell(row=r, column=col, value=data['totals'][key]).number_format = num_fmt
            ws.cell(row=r, column=col).font = bold_font

        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 16

        output = io.BytesIO()
        wb.save(output)
        return self._get_xlsx_response(output, f"aging_{report_type}_{date_to}.xlsx")

    @http.route('/ac_financial_reports/export/aging/pdf', type='http', auth='user')
    def export_aging_pdf(self, date_to=None, report_type='receivable', **kwargs):
        data = request.env['account.account'].get_aging_data(date_to=date_to, report_type=report_type)
        title = 'Aging Receivable' if report_type == 'receivable' else 'Aging Payable'
        headers = [
            {'key': 'partner_name', 'label': 'Partner'},
            {'key': 'current', 'label': 'Current', 'right': True},
            {'key': 'b1_30', 'label': '1-30', 'right': True},
            {'key': 'b31_60', 'label': '31-60', 'right': True},
            {'key': 'b61_90', 'label': '61-90', 'right': True},
            {'key': 'b90_plus', 'label': '>90', 'right': True},
            {'key': 'total', 'label': 'Total', 'right': True},
        ]
        rows = [{'_level': 0, '_css': '', **p} for p in data['partners']]
        totals = {'partner_name': 'TOTAL', **data['totals']}
        html = self._build_pdf_html(title, f"As of: {date_to}", headers, rows, totals)
        return self._get_pdf_response(html, f"aging_{report_type}_{date_to}.pdf")
